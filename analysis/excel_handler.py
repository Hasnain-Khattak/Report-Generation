import sys
import streamlit as st
import os
import tempfile
from datetime import datetime
import pandas as pd
import docx
from docx import Document
import comtypes.client
from docx.shared import Pt, Inches
from docx.enum.text import WD_PARAGRAPH_ALIGNMENT
from docx.oxml import parse_xml
from docx.oxml.ns import nsdecls
import fitz  # PyMuPDF
import io
import win32gui
import win32ui
import win32con
import win32com.client
from dotenv import load_dotenv
import pytesseract
import mammoth
import base64
import win32com.client
import pythoncom
import time
from PIL import Image
import openai
from PIL import Image, ImageDraw, ImageFont
from docx import Document
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl import Workbook, load_workbook
import re
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from docxcompose.composer import Composer
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter

# from trial import _extract_from_excel as EXTRACT_FROM_EXCEL
                

from typing import List, Dict, Tuple, Union, Optional
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)



load_dotenv(".env")

os.environ['OPENAI_API_KEY'] = os.getenv('OPENAI_API_KEY')

# Set page configuration
# st.set_page_config(
#     page_title="AI Audit Report Generator",
#     page_icon="📊",
#     layout="wide",
#     initial_sidebar_state="expanded"
# )


# Add Excel handling functions
class ExcelHandler:
    """Handle Excel operations for the Corrective Actions Register."""
    
    @staticmethod
    def create_new_register():
        """Create a new corrective actions register Excel file."""
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "CORRECTIVE ACTIONS REGISTER"
        
        # Define headers and their widths
        headers = [
            "Date", "Source of Issue", "Type", "Details", "Root Cause", 
            "Person(s) responsible for response", "Corrective Actions Implemented", "Actual close out date"
        ]
        column_widths = [12, 25, 20, 40, 25, 25, 40, 15]
        
        # Create header row
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=5, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet.column_dimensions[chr(64 + col_idx)].width = column_widths[col_idx - 1]
        
        # Add title
        sheet.merge_cells('A1:H1')
        title_cell = sheet.cell(row=1, column=1, value="CORRECTIVE ACTIONS REGISTER")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="left")
        
        # Add legend
        sheet.cell(row=3, column=5, value="Open")
        sheet.cell(row=3, column=5).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        sheet.cell(row=4, column=5, value="Corrective action in progress")
        sheet.cell(row=4, column=5).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        sheet.cell(row=5, column=5, value="Corrective action implemented")
        sheet.cell(row=5, column=5).fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        
        # Add document reference info
        sheet.cell(row=3, column=10, value="Document Reference No. PAFORM18.0")
        sheet.cell(row=4, column=10, value="Version No. 1.0")
        sheet.cell(row=5, column=10, value=f"Revision Date: {datetime.now().strftime('%d/%m/%Y')}")
        
        return workbook
    
    @staticmethod
    def add_action_to_register(excel_file, action_data):
        """Add a new corrective action to the register Excel file."""
        if excel_file:
            # Load existing file
            workbook = load_workbook(excel_file)
        else:
            # Create new file
            workbook = ExcelHandler.create_new_register()
        
        sheet = workbook.active
        
        # Find the last row with data
        last_row = 5  # Start after header row
        for row in range(6, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value is not None:
                last_row = row
            else:
                break
        
        # Add new row
        new_row = last_row + 1
        
        # Add data to the new row
        sheet.cell(row=new_row, column=1, value=action_data["Date"])
        sheet.cell(row=new_row, column=2, value=action_data["Source of Issue"])
        sheet.cell(row=new_row, column=3, value=action_data["Type"])
        sheet.cell(row=new_row, column=4, value=action_data["Details"])
        sheet.cell(row=new_row, column=5, value=action_data["Root Cause"])
        sheet.cell(row=new_row, column=6, value=action_data["Person"])
        sheet.cell(row=new_row, column=7, value=action_data["Corrective Actions Implemented"])
        sheet.cell(row=new_row, column=8, value=action_data["Actual close out date"])
        
        # Apply styling
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Color the row yellow (open) if no close date
        if not action_data["Actual close out date"]:
            for col in range(1, 9):
                sheet.cell(row=new_row, column=col).fill = yellow_fill
        
        # Apply borders and text wrapping
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
        
        for col in range(1, 9):
            cell = sheet.cell(row=new_row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        return workbook
    
    @staticmethod
    def save_register_to_bytes(workbook):
        """Convert workbook to bytes for download."""
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()
