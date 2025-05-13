import sys
import streamlit as st
import os
import tempfile
from datetime import datetime
import pandas as pd
import docx
from docx import Document
import comtypes.client
from docx.shared import Pt, Inches
from docx.enum.text import WD_PARAGRAPH_ALIGNMENT
from docx.oxml import parse_xml
from docx.oxml.ns import nsdecls
import fitz  # PyMuPDF
import io
import win32gui
import win32ui
import win32con
import win32com.client
from dotenv import load_dotenv
import pytesseract
import mammoth
import base64
import win32com.client
import pythoncom
import time
from PIL import Image
import openai
from PIL import Image, ImageDraw, ImageFont
from docx import Document
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl import Workbook, load_workbook
import re
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from docxcompose.composer import Composer
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from typing import List, Dict, Tuple, Union, Optional
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)



load_dotenv(".env")

os.environ['OPENAI_API_KEY'] = os.getenv('OPENAI_API_KEY')




# ------- Evidence Screensho tHandler Module -------
class EvidenceScreenshotHandler:
    """Class to handle capturing screenshots from evidence documents and processing them for the report."""
    
    def __init__(self):
        """Initialize the handler."""
        self.supported_formats = {
            '.pdf': self._extract_from_pdf,
            '.docx': self._extract_from_docx,
            '.doc': self._extract_from_doc,
            '.dotx': self._extract_from_docx,
            '.dot': self._extract_from_doc,
            '.xlsx': self.snap_xlsx,
            '.xls': self._extract_from_excel,
        }
        # Create temp directory for file operations
        self.temp_dir = tempfile.mkdtemp()
        # Keep track of Word COM objects to ensure proper cleanup
        self.word_app = None
        
    def process_evidence_files(self, evidence_files: List[str]) -> Dict[str, List[Dict[str, Union[str, bytes]]]]:
        """Process a list of evidence files and extract screenshots."""
        evidence_images = {}
        
        for file_path in evidence_files:
            try:
                file_ext = os.path.splitext(file_path)[1].lower()
                file_name = os.path.basename(file_path)
                
                if file_ext in self.supported_formats:
                    logger.info(f"Processing evidence file: {file_name}")
                    # Ensure valid file path
                    if not os.path.exists(file_path):
                        logger.error(f"File does not exist: {file_path}")
                        continue
                        
                    # Extract screenshots
                    # print(f"{self.supported_formats[file_ext]} calling FUNC...")
                    images = self.supported_formats[file_ext](file_path)
                    if images:
                        evidence_images[file_name] = images
                        # logger.info(f'Total images: {evidence_images.__len__()}')
                        # logger.info("Exiting")

                    #     sys.exit('Exiting...')
                    # else:
                    #     logger.info("Exiting")
                    #     sys.exit('Exiting...')

                else:
                    logger.warning(f"Unsupported file format: {file_ext} for file {file_name}")
            except Exception as e:
                logger.error(f"Error processing file {file_path}: {str(e)}")
                # Create a fallback image
                img = self._create_error_image(f"Error processing {os.path.basename(file_path)}", str(e))
                img_bytes = io.BytesIO()
                img.save(img_bytes, format="PNG")
                
                evidence_images[os.path.basename(file_path)] = [{
                    'data': img_bytes.getvalue(),
                    'format': 'png',
                    'source': os.path.basename(file_path),
                    'description': f"Error processing {os.path.basename(file_path)}"
                }]
        
        # Ensure proper cleanup
        self._cleanup_word_app()
        
        return evidence_images
    
    def _extract_from_pdf(self, file_path: str) -> List[Dict[str, Union[str, bytes]]]:
        """Extract images from PDF files."""
        images = []
        try:
            pdf_document = fitz.open(file_path)
            
            # If single page form as mentioned, just extract the first page
            # Otherwise, still handling up to 3 pages as before
            max_pages = min(1, len(pdf_document)) if len(pdf_document) >= 1 else len(pdf_document)
            
            for page_num in range(max_pages):
                page = pdf_document[page_num]
                # Increased resolution for better quality
                pix = page.get_pixmap(matrix=fitz.Matrix(2.0, 2.0))
                img_data = pix.tobytes("png")
                
                images.append({
                    'data': img_data,
                    'format': 'png',
                    'source': f"{os.path.basename(file_path)} (Page {page_num + 1})",
                    'description': f"Screenshot from {os.path.basename(file_path)}, Page {page_num + 1}"
                })
            
            pdf_document.close()
        except Exception as e:
            logger.error(f"Error extracting from PDF {file_path}: {str(e)}")
            # Create a fallback image with error message
            img = self._create_error_image(f"PDF Processing Error", str(e))
            img_bytes = io.BytesIO()
            img.save(img_bytes, format="PNG")
            
            images.append({
                'data': img_bytes.getvalue(),
                'format': 'png',
                'source': os.path.basename(file_path),
                'description': f"Error processing {os.path.basename(file_path)}"
            })
        
        return images
    
    def snap_xlsx(self, excel_path, output_png=''):
        """
        Open Excel file silently, capture data as image, and save as PNG
        Works cross-platform and optimized for stealth
        """
        # Create temp directory for intermediate files
        temp_dir = tempfile.mkdtemp()
        
        try:
            # Read the Excel file using pandas - no visible UI
            df = pd.read_excel(excel_path)
        
            # Windows-specific optimized approach
            import win32com.client as win32
            from PIL import ImageGrab
            import pythoncom
            
            # Initialize COM in thread
            pythoncom.CoInitialize()
            
            # Run Excel in background process
            excel = win32.DispatchEx("Excel.Application")
            excel.DisplayAlerts = False
            excel.Visible = 0  # Completely hidden
            excel.EnableEvents = False
            
            # Open workbook with optimized settings
            workbook = excel.Workbooks.Open(os.path.abspath(excel_path),
                                        UpdateLinks=False,
                                        ReadOnly=True)
            
            sheet = workbook.ActiveSheet
            
            # Get used range to capture exactly what's needed
            used_range = sheet.UsedRange
            used_range.CopyPicture(Format=2)  # 2 = xlBitmap
            
            # Save data to clipboard
            sheet.Application.CutCopyMode = False
            
            # Capture from clipboard
            image_buffer = io.BytesIO()
            img = ImageGrab.grabclipboard()
            
            # Close Excel without saving
            workbook.Close(SaveChanges=False)
            excel.Quit()
            
            # Release COM objects
            del sheet
            del workbook
            del excel
            pythoncom.CoUninitialize()
            
            if img:
                image_buffer = io.BytesIO()
                img.save(image_buffer, format='PNG')
                image_bytes = image_buffer.getvalue()
                # print(image_bytes)
                return [{
                    'data': image_bytes,
                    'format': 'png',
                    'source': f"{excel_path} (Sheet: I, Image {1})",
                    'description': f"Embedded image from {excel_path}, Sheet I"
                }]
                # return image_bytes
                # img.save(output_png, 'PNG')
                # return True
                
        except Exception as e:
            # Silent error handling
            return False
        finally:
            # Clean up temporary files
            import shutil
            try:
                shutil.rmtree(temp_dir)
            except:
                pass
    

    def _extract_from_excel(self, file_path: str) -> List[Dict[str, Union[str, bytes]]]:
        """Extract images from Excel files (.xlsx, .xls), including sheet snapshots."""
        images = []
        try:
            # Load the Excel workbook
            workbook = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
            file_name = os.path.basename(file_path)
            # logger.info(f"file name: {file_name}")
            
            # Iterate through all worksheets
            image_count = 0
            has_embedded_media = False
            
            # for sheet_name in workbook.sheetnames:
            #     worksheet = workbook[sheet_name]
                
            #     # Check for images in the worksheet
            #     if hasattr(worksheet, '_images') and worksheet._images:
            #         has_embedded_media = True
            #         for excel_img in worksheet._images:
            #             try:
            #                 # Access the image data
            #                 img_data = excel_img.image
            #                 # img_bytes = io.BytesIO()
            #                 # img.save(img_bytes, format="PNG")
            #                 # image_count += 1
                            
            #                 images.append({
            #                     'data': img_data.getvalue(),
            #                     'format': 'png',
            #                     'source': f"{file_name} (Sheet: {sheet_name}, Image {image_count})",
            #                     'description': f"Embedded image from {file_name}, Sheet: {sheet_name}"
            #                 })
            #             except Exception as img_e:
            #                 logger.warning(f"Error processing image in {file_name}, Sheet: {sheet_name}: {str(img_e)}")
            #                 continue
                
            #     # Optionally, check for charts (if needed)
            #     if hasattr(worksheet, '_charts') and worksheet._charts:
            #         has_embedded_media = True
            #         for chart in worksheet._charts:
            #             try:
            #                 # Convert chart to image (requires rendering, which may need external tools)
            #                 logger.info(f"Chart detected in {file_name}, Sheet: {sheet_name}. Chart export not implemented.")
            #                 # Placeholder for chart export if needed (e.g., using win32com to render Excel)
            #             except Exception as chart_e:
            #                 logger.warning(f"Error processing chart in {file_name}, Sheet: {sheet_name}: {str(chart_e)}")
            #                 continue
            
            # workbook.close()
            img_data = self.snap_xlsx(file_name)
            if img_data:
                has_embedded_media = True
            
                # append image bytes
                images.append({
                    'data': img_data,
                    'format': 'png',
                    'source': f"{file_name} (Sheet: I, Image {image_count})",
                    'description': f"Embedded image from {file_name}, Sheet I"
                })
            
            # If no images or charts were found, create snapshots of all sheets
            if not has_embedded_media:
                logger.info(f"No embedded media found in {file_name}. Creating sheet snapshots.")
                
                # Reopen the workbook since we closed it
                # workbook = openpyxl.load_workbook(file_path, read_only=False, data_only=True
                
                for sheet_name in workbook.sheetnames:
                    try:
                        # Method 1: Using pandas to convert to dataframe, then to image
                        # Read the worksheet into a pandas DataFrame
                        sheet_df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
                        
                        # Get styled HTML representation - we'll convert this to an image
                        html = sheet_df.to_html(index=False, border=1, classes='dataframe')
                        
                        # Use PIL to create an image of the table
                        # Determine dimensions based on DataFrame size
                        rows, cols = sheet_df.shape
                        img_width = max(800, 100 * cols)  # Adjust width based on columns
                        img_height = max(600, 30 * (rows + 1))  # Adjust height based on rows (+1 for header)
                        
                        # Create a blank white image
                        img = Image.new('RGB', (img_width, img_height), color='white')
                        draw = ImageDraw.Draw(img)
                        
                        # Try to load a font, use default if not available
                        try:
                            font = ImageFont.truetype("arial.ttf", 14)
                        except IOError:
                            font = ImageFont.load_default()
                        
                        # Add sheet name as title
                        draw.text((10, 10), f"Sheet: {sheet_name}", fill='black', font=font)
                        
                        # Draw column headers
                        y_offset = 40
                        for col_idx, col_name in enumerate(sheet_df.columns):
                            x_position = 10 + col_idx * 150
                            draw.text((x_position, y_offset), str(col_name), fill='black', font=font)
                        
                        # Draw data rows
                        for row_idx, row in enumerate(sheet_df.itertuples(index=False)):
                            y_position = y_offset + 30 + row_idx * 25
                            for col_idx, cell_value in enumerate(row):
                                x_position = 10 + col_idx * 150
                                draw.text((x_position, y_position), str(cell_value), fill='black', font=font)
                        
                        # Save the image to bytes
                        img_bytes = io.BytesIO()
                        img.save(img_bytes, format="PNG")
                        img_bytes.seek(0)
                        
                        images.append({
                            'data': img_bytes.getvalue(),
                            'format': 'png',
                            'source': f"{file_name} (Sheet: {sheet_name})",
                            'description': f"Snapshot of sheet {sheet_name} from {file_name}"
                        })
                        
                    except Exception as snapshot_e:
                        logger.warning(f"Error creating snapshot for sheet {sheet_name} in {file_name}: {str(snapshot_e)}")
                        # Try alternative method if first method fails
                        try:
                            # Method 2: Create simple text image with sheet data
                            # Get cell values from the worksheet
                            data = []
                            for row in worksheet.iter_rows(values_only=True):
                                data.append(row)
                            
                            # Create a simple text representation image
                            img = self._create_sheet_image(data, sheet_name, file_path)
                            img_bytes = io.BytesIO()
                            img.save(img_bytes, format="PNG")
                            
                            images.append({
                                'data': img_bytes.getvalue(),
                                'format': 'png',
                                'source': f"{file_name} (Sheet: {sheet_name})",
                                'description': f"Text snapshot of sheet {sheet_name} from {file_name}"
                            })
                        except Exception as alt_e:
                            logger.error(f"Alternative snapshot method failed for {sheet_name}: {str(alt_e)}")
                
                workbook.close()
            
            # If no images were created, provide a fallback
            if not images:
                logger.warning(f"No images or snapshots could be created for {file_name}. Creating fallback.")
                preview_text = f"Excel Preview: {file_name}\n\nNo content could be rendered as images."
                img = self._create_text_image(preview_text, file_path)
                img_bytes = io.BytesIO()
                img.save(img_bytes, format="PNG")
                
                images.append({
                    'data': img_bytes.getvalue(),
                    'format': 'png',
                    'source': file_name,
                    'description': f"Failed to create visual representation of {file_name}"
                })
                
        except Exception as e:
            logger.error(f"Error extracting from Excel {file_path}: {str(e)}")
            # Create a fallback image with error message
            img = self._create_error_image(f"Excel Processing Error", str(e))
            img_bytes = io.BytesIO()
            img.save(img_bytes, format="PNG")
            
            images.append({
                'data': img_bytes.getvalue(),
                'format': 'png',
                'source': os.path.basename(file_path),
                'description': f"Error processing {os.path.basename(file_path)}"
            })
        
        finally:
            return images

    def _create_sheet_image(self, data, sheet_name, file_path):
        """Create an image from sheet data."""
        from PIL import Image, ImageDraw, ImageFont
        
        # Determine required image size
        rows = len(data)
        cols = max(len(row) for row in data) if data else 0
        
        # Calculate the width and height of the image
        cell_width = 150
        cell_height = 30
        padding = 10
        header_height = 50
        
        width = padding * 2 + cols * cell_width
        height = padding * 2 + header_height + rows * cell_height
        
        # Create a new white image
        img = Image.new('RGB', (width, height), color='white')
        draw = ImageDraw.Draw(img)
        
        # Try to load font, use default if not available
        try:
            title_font = ImageFont.truetype("arial.ttf", 18)
            cell_font = ImageFont.truetype("arial.ttf", 14)
        except IOError:
            title_font = ImageFont.load_default()
            cell_font = ImageFont.load_default()
        
        # Draw title
        file_name = os.path.basename(file_path)
        draw.text((padding, padding), f"Excel Sheet: {sheet_name} - {file_name}", fill='black', font=title_font)
        
        # Draw grid lines
        for i in range(rows + 1):
            y = padding + header_height + i * cell_height
            draw.line([(padding, y), (width - padding, y)], fill='gray', width=1)
        
        for j in range(cols + 1):
            x = padding + j * cell_width
            draw.line([(x, padding + header_height), (x, height - padding)], fill='gray', width=1)
        
        # Draw data
        for i, row in enumerate(data):
            for j, cell in enumerate(row):
                if cell is not None:
                    x = padding + j * cell_width + 5
                    y = padding + header_height + i * cell_height + 5
                    draw.text((x, y), str(cell), fill='black', font=cell_font)
        
        return img
    
    def _cleanup_word_app(self):
        """Clean up Word application instance to prevent memory leaks."""
        try:
            if self.word_app:
                self.word_app.Quit()
                self.word_app = None
                # Give Word time to fully close
                time.sleep(1)
        except Exception as e:
            logger.error(f"Error cleaning up Word application: {str(e)}")
    
    def _initialize_word_app(self):
        """Initialize MS Word application instance."""
        try:
            # Try to reuse existing instance if possible
            if self.word_app is None:
                pythoncom.CoInitialize()
                self.word_app = win32com.client.Dispatch("Word.Application")
                self.word_app.Visible = False
                self.word_app.DisplayAlerts = False
            return True
        except Exception as e:
            logger.error(f"Failed to initialize Word: {str(e)}")
            return False
    
    def _convert_doc_to_pdf_with_comtypes(self, doc_path: str) -> Optional[str]:
        """Convert DOC/DOCX to PDF using comtypes."""
        try:
            # Define constants for Word
            wdFormatPDF = 17
            
            # Generate output path
            pdf_path = os.path.join(self.temp_dir, f"{os.path.splitext(os.path.basename(doc_path))[0]}.pdf")
            
            # Make absolute paths
            doc_path = os.path.abspath(doc_path)
            pdf_path = os.path.abspath(pdf_path)
            
            # Initialize COM
            word = comtypes.client.CreateObject('Word.Application')
            word.Visible = False
            
            try:
                # Open document
                doc = word.Documents.Open(doc_path)
                
                # Save as PDF
                doc.SaveAs(pdf_path, FileFormat=wdFormatPDF)
                doc.Close()
                
                # Check if conversion was successful
                if os.path.exists(pdf_path) and os.path.getsize(pdf_path) > 0:
                    logger.info(f"Successfully converted {doc_path} to PDF using comtypes")
                    return pdf_path
                else:
                    logger.error(f"PDF file not created or empty using comtypes")
                    return None
            finally:
                word.Quit()
        except Exception as e:
            logger.error(f"Error in comtypes conversion: {str(e)}")
            return None
    
    def _convert_docx_to_pdf(self, docx_path: str) -> Optional[str]:
        """Convert DOCX to PDF using multiple methods for better reliability."""
        # Method 1: Try with win32com
        try:
            if not self._initialize_word_app():
                logger.warning("Could not initialize Word application. Trying alternative method.")
                return self._convert_doc_to_pdf_with_comtypes(docx_path)
                
            # Generate output path
            pdf_path = os.path.join(self.temp_dir, f"{os.path.splitext(os.path.basename(docx_path))[0]}.pdf")
            
            # Make sure path exists
            os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
            
            # Ensure we have absolute path
            docx_path = os.path.abspath(docx_path)
            pdf_path = os.path.abspath(pdf_path)
            
            logger.info(f"Opening document: {docx_path}")
            doc = self.word_app.Documents.Open(docx_path)
            logger.info(f"Saving as PDF: {pdf_path}")
            
            # Try with wdFormatPDF constant
            try:
                doc.SaveAs(pdf_path, FileFormat=17)  # 17 is for PDF format
            except Exception as save_ex:
                logger.warning(f"First save attempt failed: {str(save_ex)}")
                # Try alternative method
                doc.ExportAsFixedFormat(pdf_path, 17)  # 17 = PDF
                
            doc.Close(SaveChanges=False)
            
            # Wait for the file to be fully written
            max_wait = 15  # seconds
            for _ in range(max_wait * 2):
                if os.path.exists(pdf_path) and os.path.getsize(pdf_path) > 0:
                    # Additional wait to make sure file is fully written
                    time.sleep(0.5)
                    logger.info(f"Successfully converted {docx_path} to PDF")
                    return pdf_path
                time.sleep(0.5)
            
            # If PDF wasn't created or is empty, try alternative method
            logger.warning(f"PDF conversion with win32com failed: Output file is empty or doesn't exist")
            return self._convert_doc_to_pdf_with_comtypes(docx_path)
                
        except Exception as e:
            logger.error(f"Error converting DOCX to PDF with win32com: {str(e)}")
            # Try alternative method
            return self._convert_doc_to_pdf_with_comtypes(docx_path)
    
    def _screenshot_with_direct_rendering(self, file_path: str) -> Optional[bytes]:
        """Create a direct screenshot of Word document by using visible Word application."""
        try:
            # Initialize a visible Word application for screenshot capturing
            pythoncom.CoInitialize()
            word = win32com.client.Dispatch("Word.Application")
            word.Visible = True  # Must be visible to capture screenshot
            
            try:
                # Open document
                doc = word.Documents.Open(os.path.abspath(file_path))
                
                # Maximize window and give it time to render
                word.WindowState = 1  # maximized
                word.ActiveWindow.View.Zoom.Percentage = 100  # set zoom to 100%
                time.sleep(1.5)  # Allow time for rendering
                
                # Use PIL to capture screenshot of active window
                try:
                    import pyautogui
                    screenshot = pyautogui.screenshot()
                    
                    # Convert to bytes
                    img_bytes = io.BytesIO()
                    screenshot.save(img_bytes, format="PNG")
                    img_data = img_bytes.getvalue()
                    
                    # Close document without saving changes
                    doc.Close(False)
                    return img_data
                except ImportError:
                    logger.error("pyautogui not available for direct screenshot")
                    doc.Close(False)
                    return None
            finally:
                word.Quit()
                pythoncom.CoUninitialize()
        except Exception as e:
            logger.error(f"Error in direct screenshot: {str(e)}")
            return None
    
    def _extract_from_docx(self, file_path: str) -> List[Dict[str, Union[str, bytes]]]:
        """Extract screenshot from DOCX files using multiple methods for reliability."""
        images = []
        
        # First try to convert to PDF
        pdf_path = self._convert_docx_to_pdf(file_path)
        
        if pdf_path and os.path.exists(pdf_path):
            # If conversion successful, extract from PDF
            logger.info(f"Using PDF conversion for {file_path}")
            images = self._extract_from_pdf(pdf_path)
            
            # Try to clean up the temporary PDF file
            try:
                os.remove(pdf_path)
            except:
                pass
                
            return images
        
        # If conversion failed, try direct rendering method
        logger.warning(f"PDF conversion failed for {file_path}, trying direct rendering")
        screenshot_data = self._screenshot_with_direct_rendering(file_path)
        
        if screenshot_data:
            images.append({
                'data': screenshot_data,
                'format': 'png',
                'source': os.path.basename(file_path),
                'description': f"Direct screenshot from {os.path.basename(file_path)}"
            })
            return images
        
        # If direct rendering failed, fallback to docx content extraction
        logger.warning(f"Direct rendering failed for {file_path}, using fallback method")
        
        try:
            # Open the document directly with python-docx
            doc = Document(file_path)
            
            # Get document content for preview
            preview_text = ""
            for i, para in enumerate(doc.paragraphs):
                if i < 20:  # First 20 paragraphs for preview
                    preview_text += para.text + "\n"
            
            # Create an image with the document preview
            img = self._create_text_image(preview_text, file_path)
            img_bytes = io.BytesIO()
            img.save(img_bytes, format="PNG")
            
            images.append({
                'data': img_bytes.getvalue(),
                'format': 'png',
                'source': os.path.basename(file_path),
                'description': f"Screenshot from {os.path.basename(file_path)}"
            })
            
            # Try to extract any embedded images from the document
            try:
                for rel in doc.part.rels.values():
                    if "image" in rel.target_ref:
                        image_part = rel.target_part
                        if image_part:
                            img_bytes = io.BytesIO(image_part.blob)
                            # Convert to PNG for consistency
                            img = Image.open(img_bytes)
                            output_bytes = io.BytesIO()
                            img.save(output_bytes, format="PNG")
                            
                            images.append({
                                'data': output_bytes.getvalue(),
                                'format': 'png',
                                'source': os.path.basename(file_path),
                                'description': f"Embedded image from {os.path.basename(file_path)}"
                            })
            except Exception as img_ex:
                logger.warning(f"Could not extract embedded images: {str(img_ex)}")
                
        except Exception as e:
            logger.error(f"Error extracting from DOCX {file_path}: {str(e)}")
            
            # Create a fallback image with error message
            img = self._create_error_image(f"Document Processing Error", str(e))
            img_bytes = io.BytesIO()
            img.save(img_bytes, format="PNG")
            
            images.append({
                'data': img_bytes.getvalue(),
                'format': 'png',
                'source': os.path.basename(file_path),
                'description': f"Error processing {os.path.basename(file_path)}"
            })
                    
        return images
    
    def _extract_from_doc(self, file_path: str) -> List[Dict[str, Union[str, bytes]]]:
        """Extract screenshot from DOC files using the same methods as DOCX."""
        # Since the methods are the same, reuse the docx extraction
        return self._extract_from_docx(file_path)
    
    def _create_text_image(self, text: str, file_path: str) -> Image.Image:
        """Create an image with text content from document."""
        # Create a white background image
        img = Image.new('RGB', (1000, 1200), color=(255, 255, 255))
        
        # Use PIL to add text
        draw = ImageDraw.Draw(img)
        
        # Try to get a font
        try:
            font = ImageFont.truetype("arial.ttf", 16)
            small_font = ImageFont.truetype("arial.ttf", 12)
        except IOError:
            try:
                font = ImageFont.truetype("DejaVuSans.ttf", 16)
                small_font = ImageFont.truetype("DejaVuSans.ttf", 12)
            except IOError:
                font = ImageFont.load_default()
                small_font = ImageFont.load_default()
        
        # Add file name as header
        header = f"Document Preview: {os.path.basename(file_path)}"
        draw.text((20, 20), header, fill=(0, 0, 100), font=font)
        
        # Add horizontal line
        draw.line((20, 50, 980, 50), fill=(0, 0, 0), width=2)
        
        # Add text content with word wrap
        y_position = 70
        words = text.split()
        line = ""
        for word in words:
            test_line = line + word + " "
            # Get the width of the text
            text_width = draw.textlength(test_line, font=small_font)
            if text_width < 960:
                line = test_line
            else:
                draw.text((20, y_position), line, fill=(0, 0, 0), font=small_font)
                y_position += 20
                line = word + " "
                
            # Limit the height to avoid excessive image size
            if y_position > 1100:
                draw.text((20, y_position), line + "...", fill=(0, 0, 0), font=small_font)
                y_position += 40
                draw.text((20, y_position), "[Content truncated for display]", fill=(100, 0, 0), font=small_font)
                break
        
        # Draw the last line
        if line and y_position <= 1100:
            draw.text((20, y_position), line, fill=(0, 0, 0), font=small_font)
        
        return img
    
    def _create_error_image(self, title: str, error_message: str) -> Image.Image:
        """Create an image displaying an error message."""
        # Create a white background image with red border
        img = Image.new('RGB', (800, 400), color=(255, 255, 255))
        draw = ImageDraw.Draw(img)
        
        # Draw red border
        draw.rectangle([(0, 0), (799, 399)], outline=(255, 0, 0), width=3)
        
        # Try to get a font
        try:
            title_font = ImageFont.truetype("arial.ttf", 20)
            body_font = ImageFont.truetype("arial.ttf", 14)
        except IOError:
            try:
                title_font = ImageFont.truetype("DejaVuSans.ttf", 20)
                body_font = ImageFont.truetype("DejaVuSans.ttf", 14)
            except IOError:
                title_font = ImageFont.load_default()
                body_font = ImageFont.load_default()
        
        # Add error title
        draw.text((20, 20), title, fill=(255, 0, 0), font=title_font)
        
        # Add horizontal line
        draw.line((20, 50, 780, 50), fill=(200, 0, 0), width=2)
        
        # Add error message with word wrap
        y_position = 70
        words = error_message.split()
        line = ""
        for word in words:
            test_line = line + word + " "
            # Get the width of the text
            text_width = draw.textlength(test_line, font=body_font)
            if text_width < 760:
                line = test_line
            else:
                draw.text((20, y_position), line, fill=(0, 0, 0), font=body_font)
                y_position += 20
                line = word + " "
        
        # Draw the last line
        if line:
            draw.text((20, y_position), line, fill=(0, 0, 0), font=body_font)
        
        # Add helpful message
        draw.text((20, 350), "This is a fallback image. Please check the log for more information.", 
                  fill=(100, 100, 100), font=body_font)
        
        return img
    
    @staticmethod
    def create_thumbnail(image_data: bytes, max_size: tuple = (200, 200)) -> bytes:
        """Create a thumbnail of the image."""
        try:
            img = Image.open(io.BytesIO(image_data))
            img.thumbnail(max_size, Image.LANCZOS)
            
            img_bytes = io.BytesIO()
            img.save(img_bytes, format="PNG")
            return img_bytes.getvalue()
        except Exception as e:
            logger.error(f"Error creating thumbnail: {str(e)}")
            return image_data
    
    @staticmethod
    def image_to_base64(image_data: bytes, image_format: str = 'png') -> str:
        """Convert image data to base64 for embedding in HTML."""
        try:
            base64_encoded = base64.b64encode(image_data).decode('ascii')
            return f"data:image/{image_format};base64,{base64_encoded}"
        except Exception as e:
            logger.error(f"Error converting image to base64: {str(e)}")
            return ""
            
    def clean_up(self):
        """Clean up temporary files when done."""
        # Make sure Word is closed
        self._cleanup_word_app()
        
        try:
            # Remove temp directory and its contents
            for file in os.listdir(self.temp_dir):
                try:
                    os.remove(os.path.join(self.temp_dir, file))
                except Exception as e:
                    logger.warning(f"Could not remove temp file {file}: {str(e)}")
            os.rmdir(self.temp_dir)
        except Exception as e:
            logger.error(f"Error cleaning up temporary files: {str(e)}")
